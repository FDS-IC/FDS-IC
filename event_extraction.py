import re
import settings
import json
import openpyxl
import csv
import os
import docx
from stanfordnlp_ner import StanfordNER


class EventExtraction():
    def __init__(self, context, nlp):
        self.nlp_result = nlp.ner_result
        self.news = context
        self.event = {}

        self.having_event()
        if self.event.get('trigger') in settings.FIRE_TRIGGER:
            self.fire_event()

        # 新增处理措施字段提取
        self.event['measure'] = self.extract_measure()

        print(self.event)  # 用于调试

        if self.event.get('trigger') and self.event.get('events'):
            values = [
                str(self.event.get('trigger', "")),
                str(self.event.get('events', "")),
                str(self.event.get('time', "")),
                str(self.event.get('location', "")),
                str(self.event.get('cause', "")),
                str(self.event.get('loss', "")),
                str(self.event.get('measure', ""))
            ]
            self.save_to_excel(values)
            self.save_to_csv(values)
            self.save_to_json()
            self.save_to_txt()

    def fire_event(self):
        times = self.taking_time()
        self.event['time'] = times[0] if times else "未知"
        self.event['location'] = self.taking_location()
        cause_list = pattern_match(pattern_cause(), self.news)
        self.event['cause'] = ",".join(cause_list) if cause_list else "正在进一步调查"
        lose_list = pattern_match(pattern_lose(), self.news)
        self.event['loss'] = ",".join(lose_list) if lose_list else "未知"

    def extract_measure(self):
        """正则提取‘处理措施：’后面的内容"""
        match = re.search(r"处理措施[:：](.*)", self.news)
        if match:
            return match.group(1).strip()
        return "无"

    def having_event(self):
        # 火灾类触发词判断
        for trigger_word in settings.FIRE_TRIGGER.keys():
            if trigger_word in self.news:
                self.event['trigger'] = trigger_word
                self.event['events'] = settings.FIRE_TRIGGER[trigger_word]
                return

        # 金融类触发词判断
        finance_trigger = list(settings.FINANCE_TRIGGER.keys())
        re_pattern = re.compile(r"({})".format('|'.join(map(re.escape, finance_trigger))))
        match_list = re.findall(re_pattern, self.news)

        if match_list:
            self.event['trigger'] = match_list[0]
            self.event['events'] = settings.FINANCE_TRIGGER[match_list[0]]
        else:
            self.event['trigger'] = None
            self.event['events'] = None

    def taking_time(self):
        return [text for text, tag in self.nlp_result if tag in ['DATE', 'TIME']]

    def taking_location(self):
        locations = {text for text, tag in self.nlp_result if tag in settings.LOC}
        return ",".join(sorted(locations)) if locations else "其他"

    def save_to_excel(self, event_data, filename='./output/data_1.xlsx'):
        try:
            if not os.path.exists(filename):
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.append(['trigger', 'events', 'time', 'location', 'cause', 'loss', 'measure'])
            else:
                wb = openpyxl.load_workbook(filename)
                sheet = wb.active

            row = sheet.max_row + 1
            for col, value in enumerate(event_data, 1):
                sheet.cell(row=row, column=col, value=value)

            wb.save(filename)
            print(f"✅ Excel 已保存: {filename}")
        except Exception as e:
            print(f"❌ 保存 Excel 失败: {e}")

    def save_to_csv(self, event_data, filename='./output/data_1.csv'):
        try:
            file_exists = os.path.exists(filename)
            with open(filename, mode='a', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                if not file_exists:
                    writer.writerow(['trigger', 'events', 'time', 'location', 'cause', 'loss', 'measure'])
                writer.writerow(event_data)
            print(f"✅ CSV 已保存: {filename}")
        except Exception as e:
            print(f"❌ 保存 CSV 失败: {e}")

    def save_to_json(self, filename='./output/data_1.json'):
        try:
            with open(filename, mode='a', encoding='utf-8') as f:
                json.dump(self.event, f, ensure_ascii=False)
                f.write('\n')
            print(f"✅ JSON 已保存: {filename}")
        except Exception as e:
            print(f"❌ 保存 JSON 失败: {e}")

    def save_to_txt(self, filename='./output/data_1.txt'):
        try:
            with open(filename, mode='a', encoding='utf-8') as f:
                for key, value in self.event.items():
                    f.write(f"{key}: {value}\n")
                f.write("\n")
            print(f"✅ TXT 已保存: {filename}")
        except Exception as e:
            print(f"❌ 保存 TXT 失败: {e}")


def pattern_match(patterns, text):
    result = []
    for pat_str in patterns:
        pattern = re.compile(pat_str)
        match_list = pattern.findall(text)
        if match_list:
            # match_list可能是列表，取第一个，或原样存
            if isinstance(match_list[0], tuple):
                for group in match_list[0]:
                    if group.strip():
                        result.append(group.strip())
                        break
            else:
                result.append(match_list[0].strip())
    return result


def pattern_cause():
    key_words = ['起火', '事故', '火灾', '爆炸', '目前', '泄露', '中毒', '窒息']
    patterns = [
        '.*?(?:{0})原因(.*?)[,.?:;!，。？：；！]'.format('|'.join(key_words)),
        r'(\d+事故因)'
    ]
    return patterns


def pattern_lose():
    patterns = [
        r'.*?(未造成.*?(?:伤亡|损失))[,.?:;!，。？：；！]',
        r'(\d+人死亡)',
        r'(\d+人身亡)',
        r'(\d+人受伤)',
        r'(\d+人烧伤)',
        r'(\d+人坠楼身亡)',
        r'(\d+人遇难)',
        r'(\d+被灼伤)'
    ]
    return patterns


if __name__ == '__main__':
    doc = docx.Document('燃气事故1.docx')
    for p in doc.paragraphs:
        news = p.text.strip()
        if news:
            nlp = StanfordNER(news)
            EventExtraction(news, nlp)
